import subprocess
import sys
import pkg_resources
import json
from datetime import datetime, time
import os
import asyncio
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes, JobQueue
import pytz
from openpyxl import load_workbook, Workbook

# Hàm để kiểm tra và cài đặt hoặc nâng cấp thư viện
def install_or_upgrade(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", package])

# Danh sách các thư viện cần thiết
required_packages = [
    "python-telegram-bot[job-queue]",
    "pytz",
    "openpyxl",
    "pandas",
    "nest_asyncio",
    "watchdog",
    "schedule"
]

# Kiểm tra và cài đặt các thư viện cần thiết nếu chưa có
def check_and_install_packages():
    installed_packages = {pkg.key for pkg in pkg_resources.working_set}

    for package in required_packages:
        package_name = package.split('[')[0].lower()
        if package_name not in installed_packages:
            print(f"Đang cài đặt {package}...")
            install_or_upgrade(package)

# Gọi hàm kiểm tra và cài đặt thư viện
check_and_install_packages()
# Làm sạch cửa sổ CMD
os.system('cls' if os.name == 'nt' else 'clear')
print("Tất cả các thư viện đã được cài đặt thành công.")  # In thông báo thành công

# Khởi tạo nest_asyncio để giải quyết vấn đề vòng lặp sự kiện
try:
    import nest_asyncio
except ImportError:
    print("Nest_asyncio chưa được cài đặt. Đang cài đặt...")
    install_or_upgrade('nest_asyncio')
    import nest_asyncio  # Thử lại sau khi cài đặt

nest_asyncio.apply()

# Import các thư viện cần thiết sau khi đã cài đặt
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer
from openpyxl.styles import PatternFill, Alignment
from openpyxl import load_workbook, Workbook
from telegram import Update, ChatMember
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters, JobQueue
import pandas as pd  # Đảm bảo import pandas sau khi cài đặt
import pytz
from datetime import time

# Tiếp tục với định nghĩa class và các phần còn lại của mã
class RestartBotHandler(FileSystemEventHandler):
    # Định nghĩa các phương thức trong class này
    pass

# Tạo file lưu trữ lịch sử giao dịch và số dư
data_file = 'balance_data.json'
transaction_file = 'transaction_history.json'
authorized_users_file = 'authorized_users.json'

# Hàm để load dữ liệu từ file
def load_data(file_name):
    try:
        with open(file_name, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return {}

# Hàm để lưu dữ liệu vào file
def save_data(data, file_name):
    with open(file_name, 'w') as file:
        json.dump(data, file)

# Hàm để load danh sách người dùng được ủy quyền cho một nhóm cụ thể
def load_authorized_users(chat_id):
    return load_group_data(authorized_users_file, chat_id)

# Hàm để lưu danh sách người dùng được ủy quyền cho một nhóm cụ thể
def save_authorized_users(authorized_users, chat_id):
    save_group_data(authorized_users, authorized_users_file, chat_id)

# Hàm kiểm tra xem người dùng có quyền hay không
def is_user_authorized(user_id, chat_id):
    authorized_users = load_authorized_users(chat_id)
    return str(user_id) in authorized_users

# Hàm cấp quyền cho người dùng
def authorize_user(user_id, chat_id):
    authorized_users = load_authorized_users(chat_id)
    authorized_users[str(user_id)] = True
    save_authorized_users(authorized_users, chat_id)

# Hàm hủy quyền của người dùng
def unauthorize_user(user_id, chat_id):
    authorized_users = load_authorized_users(chat_id)
    if str(user_id) in authorized_users:
        del authorized_users[str(user_id)]
    save_authorized_users(authorized_users, chat_id)

# Hàm kiểm tra xem người dùng có phải là chủ nhóm không
async def is_user_owner(chat, user_id):
    # Kiểm tra nếu cuộc trò chuyện là nhóm hoặc siêu nhóm
    if chat.type in ['group', 'supergroup']:
        admins = await chat.get_administrators()
        for admin in admins:
            if admin.user.id == user_id and admin.status == 'creator':
                return True
    # Nếu là cuộc trò chuyện riêng tư, không có quản trị viên
    elif chat.type == 'private':
        # Trong cuộc trò chuyện riêng tư, người dùng là chủ sở hữu
        return True
    return False

# Hàm kiểm tra xem tin nhắn có yêu cầu cấp quyền hay không
def is_authorization_request(text):
    return text in ["授权", "给权限", "报权限", "cấp quyền"]

# Hàm kiểm tra xem tin nhắn có yêu cầu thu hồi quyền hay không
def is_revocation_request(text):
    return text in ["撤销权限", "取消授权", "hủy quyền"]

lock = asyncio.Lock()

# Hàm để load dữ liệu từ file cho một nhóm cụ thể
def load_group_data(file_name, chat_id):
    try:
        full_file_name = f'{chat_id}_{file_name}'
        # print(f"Loading data from {full_file_name}")  # Bình luận dòng này để không in ra thông báo
        with open(full_file_name, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        # print(f"File not found: {full_file_name}")  # Bình luận dòng này để không in ra thông báo
        return {}

# Hàm để lưu dữ liệu vào file cho một nhóm cụ thể
def save_group_data(data, file_name, chat_id):
    full_file_name = f'{chat_id}_{file_name}'
    # print(f"Saving data to {full_file_name}")  # Bình luận dòng này để không in ra thông báo
    with open(full_file_name, 'w') as file:
        json.dump(data, file)

# Sử dụng các hàm này trong các hàm xử lý dữ liệu
async def update_balance_and_transaction(user_id, amount, user_name, now, chat_id):
    async with lock:
        # Cập nhật số dư
        balance_data = load_group_data(data_file, chat_id)
        if str(user_id) not in balance_data:
            balance_data[str(user_id)] = 0
        balance_data[str(user_id)] += amount
        save_group_data(balance_data, data_file, chat_id)

        # Cập nhật lịch sử giao dịch
        transaction_history = load_group_data(transaction_file, chat_id)
        if str(user_id) not in transaction_history:
            transaction_history[str(user_id)] = []
        
        # Thêm thông báo debug
        # print(f"Updating transaction for user {user_name}: {amount} at {now}")  # Bình luận dòng này
        
        transaction_history[str(user_id)].append({
            "user_name": user_name,
            "time": now,
            "amount": amount
        })
        save_group_data(transaction_history, transaction_file, chat_id)
        # print(f"Transaction history updated for chat_id: {chat_id}")  # Bình luận dòng này

# Thay thế các hàm liên quan đến việc lưu và đọc ID nhóm chat

# Biến global để lưu trữ các ID nhóm chat
group_chat_ids = set()

# Hàm để lưu ID nhóm chat
def save_group_chat_id(chat_id):
    group_chat_ids.add(chat_id)
    with open('group_chat_ids.txt', 'w') as file:
        json.dump(list(group_chat_ids), file)
    print(f"Đã lưu ID nhóm chat: {chat_id}")  # Thêm dòng này để debug

# Hàm để kiểm tra và lưu ID nhóm chat nếu chưa được lưu
def check_and_save_group_chat_id(chat_id):
    if chat_id not in group_chat_ids:
        save_group_chat_id(chat_id)

# Hàm để load các ID nhóm chat từ file khi khởi động bot
def load_group_chat_ids():
    global group_chat_ids
    try:
        with open('group_chat_ids.txt', 'r') as file:
            group_chat_ids = set(json.load(file))
        print(f"Đã load các ID nhóm chat: {group_chat_ids}")  # Thêm dòng này để debug
    except FileNotFoundError:
        group_chat_ids = set()

# Hàm để lấy tất cả ID nhóm chat đã lưu
def get_all_group_chat_ids():
    return list(group_chat_ids)

# Thêm biến global để lưu trạng thái của mỗi nhóm
group_started_status = {}

def set_group_start_status(chat_id, status):
    group_started_status[chat_id] = status

def is_group_started(chat_id):
    return group_started_status.get(chat_id, False)

# Handler để xử lý tin nhắn và lưu ID nhóm chat
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    try:
        chat_id = update.effective_chat.id
        user_id = update.message.from_user.id
        chat = update.effective_chat
 
        # Kiểm tra nếu tin nhắn là từ một cuộc trò chuyện riêng tư
        if chat.type == 'private':
            # Kiểm tra nếu người gửi không phải là chủ nhóm
            if not await is_user_owner(chat, user_id):
                # print(f"Tin nhắn từ {user_id} bị chặn vì không phải là chủ nhóm.")  # Bình luận dòng này
                return  # Không xử lý tin nhắn
 
        # Tiếp tục xử lý các tin nhắn khác
        # print(f"Handling message for chat_id: {chat_id}")  # Bình luận dòng này
        check_and_save_group_chat_id(chat_id)
        user_name = update.message.from_user.first_name
        text = update.message.text.strip()
 
        # print(f"Message from user {user_name} (ID: {user_id}): {text}")  # Bình luận dòng này
 
        # Kiểm tra nếu tin nhắn là lệnh "Bắt đầu"
        if text.lower() in ["bắt đầu", "开始"]:
            if await is_user_owner(chat, user_id) or is_user_authorized(user_id, chat_id):
                set_group_start_status(chat_id, True)
                await update.message.reply_text("设置成功：开始")
            else:
                await update.message.reply_text("您没有权限执行此操作。")
            return
 
        # Kiểm tra xem nhóm đã gửi lệnh "Bắt đầu" chưa
        if not is_group_started(chat_id):
            await update.message.reply_text(
                "状态：记账失败！\n"
                "原因：本群组今日还没有设置「开始记账」状态。\n"
                "提示：如需记账，请您发送「开始」二字。"
            )
            return
 
        # Kiểm tra xem người dùng có quyền không
        if not await is_user_owner(chat, user_id) and not is_user_authorized(user_id, chat_id):
            # Kiểm tra nếu tin nhắn là lệnh cộng tiền, trừ tiền hoặc cấp quyền
            if text.startswith('+') or text.startswith('-') or text.startswith("入款 -") or text.startswith("下发 -"):
                await update.message.reply_text("您没有权限执行此操作。")
                return
 
        # Kiểm tra nếu là câu cấp quyền
        if update.message.reply_to_message and is_authorization_request(text):
            target_user = update.message.reply_to_message.from_user
            target_user_id = target_user.id
            target_user_name = target_user.username or target_user.first_name
            if not await is_user_owner(chat, user_id) and not is_user_authorized(user_id, chat_id):
                await update.message.reply_text("您没有权限执行此操作。")
                return
            authorize_user(target_user_id, chat_id)
            await update.message.reply_text(f"已授予用户 {target_user_name} 权限。")
            return

        # Kiểm tra nếu là câu thu hồi quyền
        if update.message.reply_to_message and is_revocation_request(text):
            target_user = update.message.reply_to_message.from_user
            target_user_id = target_user.id
            target_user_name = target_user.username or target_user.first_name
            if not await is_user_owner(chat, user_id) and not is_user_authorized(user_id, chat_id):
                await update.message.reply_text("您没有权限执行此操作。")
                return
            unauthorize_user(target_user_id, chat_id)
            await update.message.reply_text(f"已撤销用户 {target_user_name} 的权限。")
            return

        # Xử lý cng hoặc trừ tiền
        now = datetime.now().strftime('%H:%M:%S')
        if text.startswith('+') or text.startswith('-'):
            try:
                amount = int(text[1:].strip())
                if text.startswith('-'):
                    amount = -amount

                # Cập nhật số dư và lịch sử giao dịch
                await update_balance_and_transaction(user_id, amount, user_name, now, chat_id)

                # Xuất file Excel
                group_chat_id = update.effective_chat.id
                excel_file_name = f'{group_chat_id}.xlsx'
                export_to_excel(group_chat_id)  # Chỉ truyền group_chat_id, không phải excel_file_name

                # Gửi báo cáo tổng kết
                await send_summary(update, chat_id)
            except ValueError:
                await update.message.reply_text("Vui lòng nhập số tiền hợp lệ.")

        elif text.startswith("入款 -") or text.startswith("下发 -"):
            parts = text.split()
            if len(parts) != 2:
                await update.message.reply_text("Định dạng không hợp lệ")
                return

            command, amount_str = parts
            amount = int(amount_str)

            if command == "入款":
                # Xử lý hủy giao dịch nạp tiền
                amount = -abs(amount)
            elif command == "下发":
                # Xử lý hủy giao dịch rút tiền (cộng ngược lại)
                amount = abs(amount)

            # Cập nhật số dư và lịch sử giao dịch
            await update_balance_and_transaction(user_id, amount, user_name, now, chat_id)

            # Xuất file Excel
            group_chat_id = update.effective_chat.id
            excel_file_name = f'{group_chat_id}.xlsx'
            export_to_excel(group_chat_id)  # Chỉ truyền group_chat_id, không phải excel_file_name

            # Gửi báo cáo tổng kết
            await send_summary(update, chat_id)

    except Exception as e:
        print(f"Lỗi khi xử lý tin nhắn: {e}")
        # Có thể thêm logging chi tiết hơn ở đây

# Lệnh cấp quyền cho thành viên
async def grant_permission(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    chat = update.effective_chat
    user_id = update.message.from_user.id

    # Chỉ cho phép chủ nhóm thực hiện
    if not await is_user_owner(chat, user_id):
        await asyncio.sleep(0.2)
        await update.message.reply_text("您没有权限执行此操作。")
        return

    if context.args:
        target_user_id = context.args[0]
        # Lấy thông tin của người dùng được cấp quyền
        target_user = await context.bot.get_chat_member(chat.id, target_user_id)
        target_user_name = target_user.user.username or target_user.user.first_name  # Lấy username hoặc tên nu không có username

        authorize_user(target_user_id, chat_id)
        await asyncio.sleep(0.2)
        await update.message.reply_text(f"已授予用户 {target_user_name} 权限。")  # Sử dụng tên tài khoản mục tiêu
    else:
        await asyncio.sleep(0.2)
        await update.message.reply_text("请指定用户。")

# Lệnh hủy quyền thành viên
async def revoke_permission(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.effective_chat.id
    chat = update.effective_chat
    user_id = update.message.from_user.id

    # Chỉ cho phép chủ nhóm thực hiện
    if not await is_user_owner(chat, user_id):
        await asyncio.sleep(0.2)
        await update.message.reply_text("您没有权限执行此操作。")
        return

    if context.args:
        target_user_id = context.args[0]
        # Lấy thông tin của người dùng được hủy quyền
        target_user = await context.bot.get_chat_member(chat.id, target_user_id)
        target_user_name = target_user.user.username or target_user.user.first_name

        unauthorize_user(target_user_id, chat_id)
        await asyncio.sleep(0.2)
        await update.message.reply_text(f"已撤销用户 {target_user_name} 的权限。")  # Sử dụng tên tài khoản mục tiêu
    else:
        await asyncio.sleep(0.2)
        await update.message.reply_text("请指定用户ID。")


# Hàm gửi báo cáo tổng kết
async def send_summary(update: Update, chat_id):
    transaction_history = load_group_data(transaction_file, chat_id)
    deposits = []
    withdrawals = []
    total_deposit = 0
    total_withdraw = 0

    # Lọc giao dịch và tạo danh sách nạp và rút
    for user_id, transactions in transaction_history.items():
        for transaction in transactions:
            if transaction['amount'] > 0:
                deposits.append(transaction)
                total_deposit += transaction['amount']
            else:
                withdrawals.append(transaction)
                total_withdraw += abs(transaction['amount'])

    # Sắp xếp lại giao dịch theo thời gian
    deposits = sorted(deposits, key=lambda x: x['time'])
    withdrawals = sorted(withdrawals, key=lambda x: x['time'])

    # Lấy 5 giao dịch nạp và rút gần nhất
    recent_deposit = deposits[-5:]
    recent_withdrawals = withdrawals[-5:]

    # Hàm để định dạng số tiền với dấu chấm và làm nổi bật
    def format_amount(amount):
        formatted = "{:,}".format(amount).replace(',', '.')
        return f'<b><code style="color:green;">{formatted}</code></b>'

    # Tạo nội dung báo cáo với số tiền có màu xanh lục, định dạng với dấu chấm và in đậm
    deposit_summary = "入款 {} 笔（显示最近 5 笔）：\n".format(len(deposits))
    for dep in recent_deposit:
        deposit_summary += f"    {dep['user_name']}    {dep['time']}    {format_amount(dep['amount'])}\n"

    withdraw_summary = "下发 {} 笔（显示最近 5 笔）：\n".format(len(withdrawals))
    for wd in recent_withdrawals:
        withdraw_summary += f"    {wd['user_name']}    {wd['time']}    {format_amount(abs(wd['amount']))}\n"

    total_summary = (
        f"总入款：{format_amount(total_deposit)}\n"
        f"总下发：{format_amount(total_withdraw)}\n"
        f"余款：{format_amount(total_deposit - total_withdraw)}"
    )

    # Gửi báo cáo vào nhóm với chế độ phân tích HTML
    await asyncio.sleep(0.2)
    await update.message.reply_text(deposit_summary + "\n" + withdraw_summary + "\n" + total_summary, parse_mode='HTML')


# Hàm để lấy ID nhóm chat từ file
def get_saved_group_chat_id():
    try:
        with open('group_chat_id.txt', 'r') as file:
            return int(file.read().strip())
    except FileNotFoundError:
        print("Không tìm thấy file group_chat_id.txt")
        return None
    except ValueError:
        print("ID nhóm chat không hợp lệ trong file")
        return None

# Hàm xuất báo cáo ra file Excel theo định dạng yêu cầu
def export_to_excel(chat_id):
    transactions = load_group_data('transaction_history.json', chat_id)

    if not transactions:
        print("No transactions found for chat_id:", chat_id)
        return

    deposit_data = []
    withdraw_data = []
    deposit_summary = {}
    withdraw_summary = {}

    for user_id, user_transactions in transactions.items():
        for transaction in user_transactions:
            amount = transaction['amount']
            user_name = transaction['user_name']
            time_str = transaction['time']

            if amount > 0:
                deposit_data.append([user_name, time_str, amount])
                deposit_summary[user_name] = deposit_summary.get(user_name, 0) + amount
            else:
                withdraw_data.append([user_name, time_str, abs(amount)])
                withdraw_summary[user_name] = withdraw_summary.get(user_name, 0) + abs(amount)

    total_deposit = sum(deposit_summary.values())
    total_withdraw = sum(withdraw_summary.values())
    balance = total_deposit - total_withdraw

    # Đặt tên file Excel theo ID nhóm chat
    excel_file_name = f'{chat_id}.xlsx'

    with pd.ExcelWriter(excel_file_name, engine='openpyxl') as writer:
        workbook = writer.book
        worksheet = workbook.create_sheet(title='Sheet1')

        today_str = datetime.now().strftime('%Y-%m-%d')
        worksheet['A1'] = today_str

        worksheet['B2'] = f'下 {len(withdraw_data)} 笔'

        for idx, row in enumerate(withdraw_data, start=4):
            worksheet[f'B{idx}'] = row[0]
            worksheet[f'C{idx}'] = row[1]
            worksheet[f'D{idx}'] = row[2]

        worksheet['F2'] = '下发分类：'

        for idx, (user, total_amount) in enumerate(withdraw_summary.items(), start=4):
            total_transactions = sum(1 for row in withdraw_data if row[0] == user)
            worksheet[f'F{idx}'] = user
            worksheet[f'G{idx}'] = f'共 {total_transactions} 笔'
            worksheet[f'H{idx}'] = total_amount

        worksheet['K2'] = f'入款 {len(deposit_data)} 笔'

        for idx, row in enumerate(deposit_data, start=4):
            worksheet[f'K{idx}'] = row[0]
            worksheet[f'L{idx}'] = row[1]
            worksheet[f'M{idx}'] = row[2]

        worksheet['O2'] = '入款分类：'

        for idx, (user, total_amount) in enumerate(deposit_summary.items(), start=4):
            total_transactions = sum(1 for row in deposit_data if row[0] == user)
            worksheet[f'O{idx}'] = user
            worksheet[f'P{idx}'] = f'共 {total_transactions} 笔'
            worksheet[f'Q{idx}'] = total_amount

        worksheet['S3'] = '总入款'
        worksheet['S4'] = '总下发'
        worksheet['S5'] = '余款'

        worksheet['T3'] = total_deposit
        worksheet['T4'] = total_withdraw
        worksheet['T5'] = balance

        # Đặt độ rộng các cột theo yêu cầu
        column_widths = {
            'A': 12, 'B': 18, 'C': 9, 'D': 9, 'F': 18, 'G': 9, 'H': 9,
            'K': 18, 'L': 9, 'M': 9, 'O': 18, 'P': 9, 'Q': 9, 'S': 9, 'T': 9
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Ẩn cột I
        worksheet.column_dimensions['I'].hidden = True

        # Merge and center các ô với màu xanh lá cây
        merge_ranges = ['B2:D2', 'F2:H2', 'K2:M2', 'O2:Q2']
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Màu xanh lá cây

        for merge_range in merge_ranges:
            worksheet.merge_cells(merge_range)
            for cell in worksheet[merge_range]:
                cell[0].fill = green_fill
                cell[0].alignment = Alignment(horizontal="center", vertical="center")

        # Tô màu vàng cho các ô S3, S4, S5, T3, T4, T5
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Màu vàng
        cells_to_color = ['S3', 'S4', 'S5', 'T3', 'T4', 'T5']
        for cell in cells_to_color:
            worksheet[cell].fill = yellow_fill
            worksheet[cell].alignment = Alignment(horizontal="center", vertical="center")

# Hàm gửi báo cáo hàng ngày
async def send_daily_report(context):
    for group_chat_id in get_all_group_chat_ids():
        excel_file_name = f'{group_chat_id}.xlsx'
        export_to_excel(group_chat_id)
        try:
            with open(excel_file_name, 'rb') as f:
                await context.bot.send_document(chat_id=group_chat_id, document=f)
            print(f"Đã gửi báo cáo cho nhóm {group_chat_id}")
        except Exception as e:
            print(f"Lỗi khi gửi báo cáo cho nhóm {group_chat_id}: {e}")

# Hàm đặt lại dữ liệu hàng ngày
async def reset_data():
    save_data({}, data_file)  
    save_data({}, transaction_file)  
    print("Dữ liệu đã được đặt lại khi có thay đổi cần thiết.")

# Lệnh bắt đầu ghi chép
async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.message.from_user.id
    chat_id = update.effective_chat.id
    user_name = update.message.from_user.first_name
    now = datetime.now().strftime('%H:%M:%S')
    await update_balance_and_transaction(user_id, 0, user_name, now, chat_id)
    await update.message.reply_text("记账已开始。请使用 +/入款 或 -/下发 来记录交易。")

# Giám sát file bot.py để tự động khởi động lại khi có thay đổi
class RestartBotHandler(FileSystemEventHandler):
    def on_modified(self, event):
        if event.src_path.endswith('bot.py'):
            print(f'{event.src_path} has been modified. Restarting bot...')
            os.execv(sys.executable, ['python'] + sys.argv)

def watch_for_changes():
    event_handler = RestartBotHandler()
    observer = Observer()
    observer.schedule(event_handler, path='.', recursive=False)
    observer.start()

    try:
        while True:
            time.sleep(0.2)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

# Định nghĩa hàm pause_daily_report
async def pause_daily_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    jobs = context.job_queue.get_jobs_by_name('daily_report')
    if jobs:
        for job in jobs:
            job.schedule_removal()
        await update.message.reply_text("Đã tạm dừng gửi báo cáo hàng ngày.")
    else:
        await update.message.reply_text("Không có báo cáo hàng ngày nào đang chạy.")

# Định nghĩa hàm resume_daily_report
async def resume_daily_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    setup_daily_job(context.application)
    await update.message.reply_text("Đã khôi phục gửi báo cáo hàng ngày.")

# Hàm để gửi báo cáo Excel
async def send_excel_report(context: ContextTypes.DEFAULT_TYPE):
    print(f"Bắt đầu gửi báo cáo Excel vào lúc {datetime.now()}")
    for group_chat_id in get_all_group_chat_ids():
        print(f"Đang xử lý nhóm {group_chat_id}")
        excel_file_name = f'{group_chat_id}.xlsx'
        export_to_excel(group_chat_id)

        try:
            # Lấy thông tin nhóm
            chat = await context.bot.get_chat(group_chat_id)
            group_name = chat.title

            with open(excel_file_name, 'rb') as file:
                await context.bot.send_document(
                    chat_id=group_chat_id, 
                    document=file,
                    caption=f"<b><code style='color:blue;'>群组: {group_name}</code></b>\n<b><code style='color:blue;'>附件: 每日统计</code></b>",  # Định dạng HTML
                    parse_mode='HTML'  # Sử dụng chế độ phân tích HTML
                )
            print(f"Đã gửi file Excel vào nhóm {group_chat_id}")
        except Exception as e:
            print(f"Lỗi khi gửi file Excel cho nhóm {group_chat_id}: {e}")
            continue

        # Xóa các file sau khi gửi
        files_to_delete = [
            f'{group_chat_id}_balance_data.json',
            f'{group_chat_id}_transaction_history.json',
            excel_file_name
        ]

        for file_name in files_to_delete:
            delete_file(file_name)

        # Đặt lại trạng thái "bắt đầu" cho nhóm này
        set_group_start_status(group_chat_id, False)
        # Gửi thông báo yêu cầu lệnh "bắt đầu"
        await context.bot.send_message(chat_id=group_chat_id, 
                                       text='请发送"开始"或"bắt đầu"命令以开始新的记录。\nVui lòng gửi lệnh "bắt đầu" hoặc "开始" để bắt đầu ghi chép mới.')

    print(f"Đã gửi xong báo cáo Excel và xóa các file vào lúc {datetime.now()}")

# Định nghĩa hàm delete_file
def delete_file(file_name):
    try:
        if os.path.exists(file_name):
            os.remove(file_name)
            print(f"Đã xóa file {file_name}")
        else:
            print(f"File {file_name} không tồn tại")
    except PermissionError as e:
        print(f"Không thể xóa file {file_name}: {e}")

# Định nghĩa hàm setup_daily_job
def setup_daily_job(application):
    job_queue = application.job_queue
    if job_queue is None:
        print("Job queue is not initialized.")
        return

    vietnam_tz = pytz.timezone('Asia/Ho_Chi_Minh')
    report_time = time(hour=1, minute=31, tzinfo=vietnam_tz)

    job_queue.run_daily(send_excel_report, time=report_time, name='daily_report')

# Thêm hàm mới để xử lý lệnh help
async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat = update.effective_chat
    user_id = update.message.from_user.id

    # Kiểm tra nếu tin nhắn là từ một cuộc trò chuyện riêng tư
    if chat.type == 'private':
        # Kiểm tra nếu người dùng là chủ nhóm
        # Lấy danh sách các nhóm mà người dùng là chủ
        group_chat_ids = get_all_group_chat_ids()
        is_owner = False
        for group_chat_id in group_chat_ids:
            chat = await context.bot.get_chat(group_chat_id)
            if await is_user_owner(chat, user_id):
                is_owner = True
                break

        if not is_owner:
            return  # Không trả lời nếu không phải là chủ nhóm

    user_name = update.message.from_user.first_name
    help_text = f"""
    用户 {user_name} 查询的机器人使用方法如下：

    ------------------------

    【注意】

    每一天，您需要在群内发送"开始"才能开始记账。记账数据将在越南时间晚上11点重置。

    ------------------------

    【入款和下发】

    例如在群内发送：

    入款 1000
    代表入款统计中增加 1000 的金额。

    下发 2000
    代表下发 2000 的金额。

    您也可以省略「入款」和「下发」的汉字，使用加号「+」代表「入款」，使用减号「-」代表「下发」，例如在群组中发送：

    +3000
    代表入款统计中增加 3000 的金额。

    -4000
    代表下发 4000 的金额。

    ------------------------

    【撤销记账】

    在撤销记账时，必须在消息中指明是哪种操作，不可以使用加减号缩略记账的方法，操作示例如下，在群组中发送：

    入款 -5000
    代表入款统计中撤销 5000 的金额。

    下发 -6000
    代表下发统计中撤销 6000 的金额。

    ------------------------
        
    【授权人员】

    请以「授权」二字回复某个人员在群里的发言，这样可以将机器人记账权限授权给该人员，令其有权限进行记账。注意您需要将该命令在群组中以「回复某人消息」的方式发送：

    授权

    ------------------------

    【取消授权】

    请以「取消授权」四字回复某个人员在群里的发言，这样可以取消某人的记账权限。注意您需要将该命令在群组中以「回复某人消息」的方式发送：

    取消授权
    """
    await update.message.reply_text(help_text)

# Định nghĩa hàm handle_new_member
async def handle_new_member(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    for member in update.message.new_chat_members:
        if member.id == context.bot.id:
            chat_id = update.effective_chat.id
            print(f"Bot được thêm vào nhóm mới: {chat_id}")
            check_and_save_group_chat_id(chat_id)
            await update.message.reply_text("你好！我已准备好开始记录。请发送 '开始' 或 'bắt đầu' 命令以开始。")

# Hàm main
def main():
    TOKEN = "7941025829:AAEJmUDK3Jr5cbNpNjeTxIim35LZEmWr2DY"
    application = ApplicationBuilder().token(TOKEN).build()

    application.add_handler(CommandHandler("start", start_command))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    application.add_handler(CommandHandler("grant", grant_permission))
    application.add_handler(CommandHandler("revoke", revoke_permission))
    application.add_handler(CommandHandler("pause_report", pause_daily_report))
    application.add_handler(CommandHandler("resume_report", resume_daily_report))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(MessageHandler(filters.Regex(r'^帮助$'), help_command))
    application.add_handler(MessageHandler(filters.StatusUpdate.NEW_CHAT_MEMBERS, handle_new_member))

    setup_daily_job(application)

    application.run_polling()

if __name__ == "__main__":
    main()

print(f"Current working directory: {os.getcwd()}")
print(f"Directory contents: {os.listdir()}")

# Thêm đoạn code này vào cuối file hoặc trong hàm main
if os.path.exists('group_chat_ids.txt'):
    with open('group_chat_ids.txt', 'r') as file:
        content = file.read()
        print(f"Nội dung file group_chat_ids.txt: {content}")
else:
    print("File group_chat_ids.txt không tồn tại")

